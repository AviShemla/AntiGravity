import os
import pandas as pd
import glob
from openpyxl import load_workbook

BASE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'financial_data')

TARGET_ETFS = ['XLK', 'XLV', 'XLY', 'XLF', 'XLC', 'XLI', 'XLE', 'XLP', 'XLU', 'XLRE', 'XLB']

for etf in TARGET_ETFS:
    file_path = os.path.join(BASE_DIR, f'{etf}_Bayesian_Scorecard.xlsx')
    if os.path.exists(file_path):
        # We need to preserve the formatting, so we use openpyxl
        wb = load_workbook(file_path)
        if etf in wb.sheetnames:
            ws = wb[etf]
            # Find the last row
            last_row = ws.max_row
            # Check if the last row's "Model Hit" is "Pending" (usually column K or L)
            # The 'Date' column is column A (1)
            # Find which column is Date
            date_col = 1
            ws.cell(row=last_row, column=date_col).value = '2026-06-16'
            wb.save(file_path)
            print(f"Patched {etf}")
        else:
            print(f"Sheet {etf} not found in {file_path}")

# Now recompile the master scorecard
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import subprocess
subprocess.run([r"C:\Users\AviShemla\AppData\Local\Python\pythoncore-3.14-64\python.exe", os.path.join(os.path.dirname(os.path.abspath(__file__)), "compile_etf_scorecards.py")], cwd=os.path.dirname(os.path.abspath(__file__)), check=True)
print("Master scorecard recompiled.")
