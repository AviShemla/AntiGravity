import os
import pandas as pd
import numpy as np
import openpyxl
from failover_downloader import log_warning, safe_print
print = safe_print

BASE_DIR = r'C:\Users\AviShemla\AntiGravity\financial_data'
STOCK_EXCEL = os.path.join(BASE_DIR, 'Top5_Bayesian_Scorecard_Formatted.xlsx')
ETF_EXCEL = os.path.join(BASE_DIR, 'All_ETFs_Scorecard.xlsx')

def check_scorecard_bounds(excel_path, file_desc):
    if not os.path.exists(excel_path):
        print(f"[QA CHECK] {file_desc} not found. Skipping QA.")
        return
        
    print(f"\n[QA CHECK] Running bounds checking on {file_desc} at {excel_path}...")
    xls = pd.ExcelFile(excel_path)
    
    anomalies_found = False
    
    # We will use openpyxl to edit the file in place if anomalies are found
    wb = openpyxl.load_workbook(excel_path)
    
    for sheet in xls.sheet_names:
        # Read the sheet data (skiprows=2 to match our formatting)
        df = pd.read_excel(xls, sheet_name=sheet, skiprows=2)
        if df.empty:
            continue
            
        last_row = df.iloc[-1]
        
        # Determine column names based on stock vs ETF headers
        prob_col = None
        ret_col = None
        rec_col = None
        override_col = None
        
        for col in df.columns:
            if 'probability' in col.lower() or 'p(up)' in col.lower():
                prob_col = col
            if 'expected return' in col.lower() or 'expected return %' in col.lower():
                ret_col = col
            if 'recommendation' in col.lower():
                rec_col = col
            if 'override' in col.lower():
                override_col = col
                
        if prob_col is None or ret_col is None:
            print(f"  [QA WARNING] Sheet {sheet} is missing probability or return columns. Skipping bounds check.")
            continue
            
        prob = float(last_row[prob_col])
        ret = float(last_row[ret_col])
        
        # Check bounds
        prob_invalid = not (0.0 <= prob <= 1.0)
        # Expected daily returns should not be larger than 20% or smaller than -20% (expressed as decimal or percentage)
        ret_val = ret * 100.0 if abs(ret) <= 1.0 else ret # adjust if decimal vs percentage
        ret_invalid = not (-20.0 <= ret_val <= 20.0)
        
        if prob_invalid or ret_invalid:
            anomalies_found = True
            err_msg = f"🚨 MODEL ANOMALY DETECTED: Ticker {sheet} has out-of-bounds metrics (P(UP)={prob*100:.1f}%, Expected Return={ret_val:.2f}%)."
            print(f"  [QA FAIL] {err_msg} Quarantine triggered.")
            log_warning(err_msg)
            
            # Edit the cell values in openpyxl worksheet
            ws = wb[sheet]
            # Headers are on row 3 (index 2 in pandas is row 3 in Excel, so data rows start on row 4)
            # Find the row index of the last row
            last_excel_row = len(df) + 3 # row index 0-indexed + skiprows=2 + 1-indexed Excel row
            
            # Find the column index for recommendation and override note
            # Let's search columns in Excel row 3
            headers = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
            
            rec_idx = None
            override_idx = None
            kelly_idx = None
            
            for col_idx, h in enumerate(headers, 1):
                if h and 'recommendation' in str(h).lower():
                    rec_idx = col_idx
                if h and 'override' in str(h).lower():
                    override_idx = col_idx
                if h and 'kelly' in str(h).lower():
                    kelly_idx = col_idx
                    
            if rec_idx:
                ws.cell(row=last_excel_row, column=rec_idx, value="Hold")
            if override_idx:
                ws.cell(row=last_excel_row, column=override_idx, value="HOLD: QA Quarantine (Bounds Violation)")
            if kelly_idx:
                ws.cell(row=last_excel_row, column=kelly_idx, value=0.0)
                
    if anomalies_found:
        wb.save(excel_path)
        print(f"[QA SUMMARY] Bounds violation resolved. Scorecard edited and saved.")
    else:
        print(f"[QA SUMMARY] {file_desc} bounds validation passed.")

def run_qa_checks():
    # 1. Check stock scorecard
    check_scorecard_bounds(STOCK_EXCEL, "Stock Bayesian Scorecard")
    # 2. Check compiled ETF scorecard
    check_scorecard_bounds(ETF_EXCEL, "Compiled ETF Bayesian Scorecard")
    # 3. Check individual ETF scorecards
    for etf in ["XLK", "XLV", "XLY", "XLF", "XLC", "XLI", "XLE", "XLP", "XLU", "XLRE", "XLB"]:
        etf_path = os.path.join(BASE_DIR, f'{etf}_Bayesian_Scorecard.xlsx')
        check_scorecard_bounds(etf_path, f"Individual ETF Scorecard ({etf})")

if __name__ == '__main__':
    run_qa_checks()
