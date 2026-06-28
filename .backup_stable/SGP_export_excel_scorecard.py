# =============================================================================
# SGP SCORECARD EXCEL EXPORTER
# Generates a formatted Excel scorecard matching the target structure:
#   - Header rows: Ticker, Model, Predictors
#   - Columns: date(lag1), date(lag2), actual date,
#              predicted value lag1, predicted value lag2, actual return,
#              predicted direction lag1, predicted direction lag2, actual direction,
#              recommendation lag1, hit lag1, recommendation lag2, hit lag2
# =============================================================================
import pandas as pd
import numpy as np
from sklearn.cluster import KMeans
import pymc as pm
import os
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter

os.environ["PYTENSOR_FLAGS"] = "cxx="

input_file      = r'C:\Users\AviShemla\AntiGravity\financial_data\SP500_Clean_Advanced_Analysis.csv'
champion_ticker = 'JPM'
N_ROWS          = 30   # number of most-recent rows to show in the scorecard

print("Reading data...")
df = pd.read_csv(input_file)
df['Date'] = pd.to_datetime(df['Date'])
df = df.drop_duplicates(subset=['Date', 'Ticker']).sort_values(['Ticker', 'Date'])

df['RAS_Signal_Num'] = df['RAS_Signal'].map({'BUY': 1, 'HOLD': 0, 'SELL': -1}).fillna(0)
df['Market_Fear_Level_Num'] = df['Market_Fear_Level'].map({'Complacency / Calm': 0, 'High Volatility': 1}).fillna(0)

return_pivot   = df.pivot(index='Date', columns='Ticker', values='Daily_Return_%')
stdev_pivot    = df.pivot(index='Date', columns='Ticker', values='Daily_STDEV')
rsi_pivot      = df.pivot(index='Date', columns='Ticker', values='RSI_14d')
adx_pivot      = df.pivot(index='Date', columns='Ticker', values='ADX_14d')
plus_di_pivot  = df.pivot(index='Date', columns='Ticker', values='Plus_DI_14d')
minus_di_pivot = df.pivot(index='Date', columns='Ticker', values='Minus_DI_14d')
atr_pivot      = df.pivot(index='Date', columns='Ticker', values='ATR_14d')
ras_pivot      = df.pivot(index='Date', columns='Ticker', values='RAS_Signal_Num')

std_adj_returns = return_pivot / (stdev_pivot + 1e-8)

predictors_list = []
for ticker in std_adj_returns.columns:
    predictors_list.append(pd.DataFrame({
        f'{ticker}_RET_ADJ':  std_adj_returns[ticker],
        f'{ticker}_RSI':      rsi_pivot[ticker],
        f'{ticker}_ADX':      adx_pivot[ticker],
        f'{ticker}_PLUS_DI':  plus_di_pivot[ticker],
        f'{ticker}_MINUS_DI': minus_di_pivot[ticker],
        f'{ticker}_ATR':      atr_pivot[ticker],
        f'{ticker}_RAS':      ras_pivot[ticker],
    }))
all_predictors_df = pd.concat(predictors_list, axis=1)
macro_df = df.drop_duplicates(subset=['Date']).set_index('Date')[['VIX_Close', 'Market_Fear_Level_Num']]
all_predictors_df = pd.concat([all_predictors_df, macro_df], axis=1)


def run_sgp(target_series, predictors_df, lag):
    target = target_series.copy()
    shifted = predictors_df.shift(lag)
    data = pd.concat([target, shifted], axis=1).replace([np.inf, -np.inf], np.nan).dropna()

    y_full = data.iloc[:, 0].values
    X_pool = data.iloc[:, 1:]
    split  = int(len(data) * 0.8)

    y_train, y_test       = y_full[:split], y_full[split:]
    Xp_train, Xp_test     = X_pool.iloc[:split], X_pool.iloc[split:]

    corrs  = Xp_train.corrwith(pd.Series(y_train, index=Xp_train.index))
    top7   = corrs.abs().sort_values(ascending=False).head(7).index.tolist()

    Xt = Xp_train[top7].values;  Xe = Xp_test[top7].values
    Xm = Xt.mean(0);             Xs = Xt.std(0) + 1e-8
    Xt_s = (Xt - Xm) / Xs;      Xe_s = (Xe - Xm) / Xs

    with pm.Model():
        ls  = pm.Gamma("ls", alpha=2, beta=1, shape=7)
        eta = pm.HalfNormal("eta", sigma=float(np.std(y_train)))
        cov = eta**2 * pm.gp.cov.ExpQuad(input_dim=7, ls=ls)
        Xu  = KMeans(n_clusters=min(150, len(Xt_s)), random_state=42, n_init=5).fit(Xt_s).cluster_centers_
        gp  = pm.gp.MarginalApprox(cov_func=cov, approx="FITC")
        sig = pm.HalfNormal("sigma", sigma=float(np.std(y_train)))
        _   = gp.marginal_likelihood("y_obs", X=Xt_s, Xu=Xu, y=y_train, sigma=sig)
        tr  = pm.sample(draws=300, tune=300, chains=2, target_accept=0.9,
                        random_seed=42, progressbar=False)
        fp  = gp.conditional("f_pred", Xnew=Xe_s)
        pp  = pm.sample_posterior_predictive(tr, var_names=["f_pred"], progressbar=False)

    y_pred = pp.posterior_predictive["f_pred"].mean(dim=["chain", "draw"]).values
    return data.index[split:], y_test, y_pred, top7


# ---- Run both lags ----
target_adj = std_adj_returns[champion_ticker].rolling(3).mean()

print("\nRunning SGP Lag 1...")
dates1, ytest1, ypred1, top7_lag1 = run_sgp(target_adj, all_predictors_df, 1)
print("\nRunning SGP Lag 2...")
dates2, ytest2, ypred2, top7_lag2 = run_sgp(target_adj, all_predictors_df, 2)

# ---- Build aligned scorecard ----
df1 = pd.DataFrame({'Actual_AdjMA3': ytest1, 'Pred_Lag1': ypred1}, index=dates1)
df2 = pd.DataFrame({'Pred_Lag2': ypred2}, index=dates2)
raw = return_pivot[champion_ticker]

sc = df1.join(df2, how='inner').join(raw.rename('Actual_Return'), how='left')

# Business-day lag dates (shift index back)
biz_dates = pd.bdate_range(start=sc.index.min() - pd.Timedelta(days=10),
                            end=sc.index.max())
sc['date_lag1'] = [biz_dates[biz_dates < d][-1] if len(biz_dates[biz_dates < d]) else pd.NaT for d in sc.index]
sc['date_lag2'] = [biz_dates[biz_dates < d][-2] if len(biz_dates[biz_dates < d]) >= 2 else pd.NaT for d in sc.index]

# Direction of MODEL predictions (based on z-score sign)
sc['Dir_Lag1']   = np.where(sc['Pred_Lag1']   > 0, 'UP', 'Down')
sc['Dir_Lag2']   = np.where(sc['Pred_Lag2']   > 0, 'UP', 'Down')

# BUG FIX: Actual direction MUST use the raw daily return (column F),
# NOT the MA3 z-score — these can disagree in sign on any given day.
sc['Dir_Actual'] = np.where(sc['Actual_Return'] > 0, 'UP', 'Down')

sc['Rec_Lag1'] = np.where(sc['Pred_Lag1'] > 0, 'Buy', 'Sell')
sc['Rec_Lag2'] = np.where(sc['Pred_Lag2'] > 0, 'Buy', 'Sell')

# Hit = model direction matches ACTUAL RAW RETURN direction
sc['Hit_Lag1'] = np.where(sc['Dir_Lag1'] == sc['Dir_Actual'], 'On target', 'Miss')
sc['Hit_Lag2'] = np.where(sc['Dir_Lag2'] == sc['Dir_Actual'], 'On target', 'Miss')

last = sc.tail(N_ROWS).copy()

# ===== Build Excel workbook =====
wb = openpyxl.Workbook()
ws = wb.active
ws.title = f"{champion_ticker} SGP Scorecard"

# ---- Colour palette ----
BLUE_DARK  = PatternFill("solid", fgColor="1F3864")
BLUE_MID   = PatternFill("solid", fgColor="2E75B6")
BLUE_LIGHT = PatternFill("solid", fgColor="D6E4F0")
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
RED_FILL   = PatternFill("solid", fgColor="FCE4D6")
GREY_FILL  = PatternFill("solid", fgColor="F2F2F2")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

WHITE_BOLD  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DARK_BOLD   = Font(name="Calibri", bold=True, color="1F3864", size=10)
DARK_NORMAL = Font(name="Calibri", color="1F3864", size=10)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style(cell, fill=None, font=None, align=None, border=None, num_fmt=None):
    if fill:   cell.fill   = fill
    if font:   cell.font   = font
    if align:  cell.alignment = align
    if border: cell.border = border
    if num_fmt: cell.number_format = num_fmt

# ---- Row 1: Ticker header ----
ws.merge_cells("A1:M1")
c = ws["A1"]
c.value = f"Ticker Predicted: {champion_ticker}"
style(c, fill=BLUE_DARK, font=WHITE_BOLD, align=LEFT)
ws.row_dimensions[1].height = 22

# ---- Row 2: Model info ----
ws.merge_cells("A2:C2")
c = ws["A2"]
c.value = "Model chosen: SGP (Sparse Gaussian Process)"
style(c, fill=BLUE_MID, font=WHITE_BOLD, align=LEFT)

ws.merge_cells("D2:G2")
c = ws["D2"]
c.value = f"Predictors (Lag1): {', '.join(top7_lag1)}"
style(c, fill=BLUE_MID, font=WHITE_BOLD, align=LEFT)

ws.merge_cells("H2:M2")
c = ws["H2"]
c.value = f"Predictors (Lag2): {', '.join(top7_lag2)}"
style(c, fill=BLUE_MID, font=WHITE_BOLD, align=LEFT)
ws.row_dimensions[2].height = 22

# ---- Row 3: Column headers ----
headers = [
    ("A3", "date\n(lag1)"),
    ("B3", "date\n(lag2)"),
    ("C3", "date"),
    ("D3", "model predicted value\n(z-score)\nusing Lag1"),
    ("E3", "model predicted value\n(z-score)\nusing Lag2"),
    ("F3", "actual value\ndaily return %"),
    ("G3", "model predicted\ndirection daily return\nusing Lag1"),
    ("H3", "model predicted\nDirection daily return\nusing Lag2"),
    ("I3", "actual Direction\ndaily return"),
    ("J3", "recommendation\nbased on lag1\n(e.g. \"BUY\", \"Sell\", \"Hold\")"),
    ("K3", "model hit IND\nlag1"),
    ("L3", "recommendation\nbased on lag2\n(e.g. \"BUY\", \"Sell\", \"Hold\")"),
    ("M3", "model hit IND\nlag2"),
]
for addr, txt in headers:
    c = ws[addr]
    c.value = txt
    style(c, fill=BLUE_DARK, font=WHITE_BOLD, align=CENTER, border=THIN_BORDER)
ws.row_dimensions[3].height = 60

# ---- Data rows ----
for i, (idx, row) in enumerate(last.iterrows()):
    r = i + 4
    row_fill = WHITE_FILL if i % 2 == 0 else GREY_FILL

    def w(col, val, fmt=None, fill=None):
        c = ws.cell(row=r, column=col, value=val)
        style(c, fill=fill or row_fill, font=DARK_NORMAL, align=CENTER, border=THIN_BORDER, num_fmt=fmt)
        return c

    w(1, row['date_lag1'].strftime('%d/%m/%Y') if pd.notna(row['date_lag1']) else "")
    w(2, row['date_lag2'].strftime('%d/%m/%Y') if pd.notna(row['date_lag2']) else "")
    w(3, idx.strftime('%d/%m/%Y'))
    # Predicted values are z-scores (return/stdev), NOT raw %. Show as number.
    w(4, round(row['Pred_Lag1'], 4),    fmt='0.0000')
    w(5, round(row['Pred_Lag2'], 4),    fmt='0.0000')
    # Actual return is raw % — show as percentage
    w(6, round(row['Actual_Return']/100, 4), fmt='0.00%')
    w(7, row['Dir_Lag1'])
    w(8, row['Dir_Lag2'])
    w(9, row['Dir_Actual'])

    # Recommendation col J — colour code
    rec1_fill = GREEN_FILL if row['Rec_Lag1'] == 'Buy' else RED_FILL
    w(10, row['Rec_Lag1'], fill=rec1_fill)

    # Hit indicator col K
    hit1_fill = GREEN_FILL if row['Hit_Lag1'] == 'On target' else RED_FILL
    w(11, row['Hit_Lag1'], fill=hit1_fill)

    # Recommendation col L
    rec2_fill = GREEN_FILL if row['Rec_Lag2'] == 'Buy' else RED_FILL
    w(12, row['Rec_Lag2'], fill=rec2_fill)

    # Hit indicator col M
    hit2_fill = GREEN_FILL if row['Hit_Lag2'] == 'On target' else RED_FILL
    w(13, row['Hit_Lag2'], fill=hit2_fill)

    ws.row_dimensions[r].height = 18

# ---- Summary row ----
sr = len(last) + 4
ws.merge_cells(f"A{sr}:C{sr}")
c = ws[f"A{sr}"]
c.value = "Directional Accuracy"
style(c, fill=BLUE_DARK, font=WHITE_BOLD, align=CENTER, border=THIN_BORDER)

lag1_acc = (last['Hit_Lag1'] == 'On target').mean()
lag2_acc = (last['Hit_Lag2'] == 'On target').mean()

ws.merge_cells(f"D{sr}:G{sr}")
c = ws[f"D{sr}"]
c.value = f"Lag1: {lag1_acc:.1%}"
style(c, fill=BLUE_MID, font=WHITE_BOLD, align=CENTER, border=THIN_BORDER)

ws.merge_cells(f"H{sr}:M{sr}")
c = ws[f"H{sr}"]
c.value = f"Lag2: {lag2_acc:.1%}"
style(c, fill=BLUE_MID, font=WHITE_BOLD, align=CENTER, border=THIN_BORDER)
ws.row_dimensions[sr].height = 22

# ---- Column widths ----
col_widths = [13, 13, 13, 16, 16, 14, 14, 14, 12, 22, 14, 22, 14]
for ci, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

# ---- Freeze panes ----
ws.freeze_panes = "A4"

out_path = r'C:\Users\AviShemla\AntiGravity\financial_data\JPM_SGP_Scorecard.xlsx'
wb.save(out_path)
print(f"\nExcel scorecard saved to: {out_path}")
print(f"Lag1 accuracy (last {N_ROWS} days): {lag1_acc:.1%}")
print(f"Lag2 accuracy (last {N_ROWS} days): {lag2_acc:.1%}")
